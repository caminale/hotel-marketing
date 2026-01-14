import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils.dataframe import dataframe_to_rows

print("="*80)
print("🎯 CRÉATION DU TABLEAU DE BORD - DÉPENSES 2025")
print("="*80)

# ============================================================================
# ÉTAPE 1: CHARGEMENT ET NETTOYAGE DES DONNÉES
# ============================================================================
print("\n📂 Chargement des données...")
fichier_source = "Grand Hôtel des Bains Export 2025.xlsx"
df = pd.read_excel(fichier_source, header=2)

# Renommer les colonnes
df = df.rename(columns={
    'Unnamed: 0': 'Type_operation',
    'Unnamed: 1': 'Categorie_detaillee',
    'Unnamed: 2': 'Categorie',
    'Unnamed: 3': 'Intitule',
    'Unnamed: 4': 'Montant_TTC',
    'Unnamed: 5': 'Date_reglement',
    'Unnamed: 6': 'Date_facturation'
})

# Nettoyer et filtrer - SUPPRIMER LES ENCAISSEMENTS
df = df[df['Type_operation'] != 'Type d\'opération']
df = df[df['Type_operation'] == 'Décaissement'].copy()
df['Montant_TTC'] = pd.to_numeric(df['Montant_TTC'], errors='coerce')
df = df[df['Montant_TTC'].notna()]

print(f"   ✓ {len(df)} dépenses chargées (encaissements exclus)")

# ============================================================================
# DICTIONNAIRE DES FOURNISSEURS CONNUS
# ============================================================================
FOURNISSEURS_CONNUS = {
    # Logiciels & Services
    'HOTELIOS': 'HOTELIOS (Logiciel hôtelier)',
    'IWINERESTO': 'IWINERESTO (Logiciel restauration)',
    'CEGID': 'CEGID (Logiciel comptable)',
    'BOOKING.COM': 'BOOKING.COM',
    'BOOKING': 'BOOKING.COM',
    
    # Télécom
    'ORANGE': 'ORANGE',
    'FREE': 'FREE',
    'FREE PRO': 'FREE',
    
    # Assurances
    'AXA': 'AXA (Assurance)',
    
    # Charges sociales
    'URSSAF': 'URSSAF',
    'DGFIP': 'DGFIP (Impôts)',
    
    # Banque & Prêts
    'BPIFRANCE': 'BPI FRANCE',
    
    # Publicité
    'SOLOCAL': 'SOLOCAL (Pages Jaunes)',
    'CHARMHOTEL': 'CHARMHOTEL',
    'GIE CHARMHOTEL': 'CHARMHOTEL',
    
    # BTP & Travaux
    'MFG': 'MFG (Honoraire BTP)',
    'SUDEL': 'SUDEL (BTP)',
    
    # Eau
    'REGIE SYNDICAT': 'Régie Syndicat (Eau)',
    
    # Énergie
    'EDF': 'EDF',
    'EDF SA': 'EDF',
    
    # Blanchisserie
    'SDEZ': 'SDEZ (Blanchisserie)',
    
    # Grossistes alimentaires
    'RELAIS D OR': "RELAIS D'OR",
    'DROMADIS': 'DROMADIS',
    'LA NATURE': 'LA NATURE À TABLE',
    
    # Divers
    'XEROX': 'XEROX',
    'PREFILOC': 'PREFILOC (Location)',
    'SPRE': 'SPRE (SACEM)',
    'SACEM': 'SACEM',
    'POLYFROID': 'POLYFROID (Climatisation)',
}

# ============================================================================
# FONCTION D'EXTRACTION DES FOURNISSEURS (AMÉLIORÉE)
# ============================================================================
def extraire_fournisseur(intitule):
    """Extrait le vrai nom du fournisseur de l'intitulé"""
    if pd.isna(intitule):
        return "Non spécifié"
    
    intitule = str(intitule).strip()
    intitule_upper = intitule.upper()
    
    # ========== CAS SPÉCIAUX ==========
    
    # PRLV INT ETS DGFIP - Prélèvements impôts
    if 'PRLV INT' in intitule_upper and 'DGFIP' in intitule_upper:
        return "DGFIP (Impôts)"
    
    # VIREMENT EMIS WEB - extraire le nom après
    if 'VIREMENT EMIS' in intitule_upper:
        match = re.search(r'VIREMENT EMIS\s+(?:WEB\s+)?([A-Za-z][A-Za-z\s]+)', intitule, re.IGNORECASE)
        if match:
            fournisseur = match.group(1).strip()
            for key, value in FOURNISSEURS_CONNUS.items():
                if key.upper() in fournisseur.upper():
                    return value
            return fournisseur[:35]
    
    # Salaires saisonniers (VIR SEPA :20 saison, :21 saison, etc.)
    if re.search(r':\d+\s+saison', intitule, re.IGNORECASE):
        return "Salaires saisonniers"
    
    # Paiements directs connus
    if 'AMAZON' in intitule_upper:
        return "AMAZON"
    if 'GOOGLE' in intitule_upper and 'ADS' in intitule_upper:
        return "GOOGLE ADS"
    if 'LA POSTE' in intitule_upper:
        return "LA POSTE"
    if 'CANVA' in intitule_upper:
        return "CANVA"
    
    # ========== VIR SEPA ==========
    if 'VIR SEPA' in intitule:
        # Pattern: VIR SEPA [NOM DU FOURNISSEUR] [reste...]
        match = re.search(r'VIR SEPA\s+([A-Za-z][A-Za-z0-9\s\.\'\-&]+?)(?:\s+(?:facture|solde|acompte|P\d|F\d|DEVIS|\d{4,}|FACTURE|Salaire|Virement|pour|avance|reboursement|COTISATION)|\s+-\s+|\s*$)', intitule, re.IGNORECASE)
        if match:
            fournisseur = match.group(1).strip()
            fournisseur = re.sub(r'\s+', ' ', fournisseur).strip()
            # Vérifier dans le dictionnaire
            for key, value in FOURNISSEURS_CONNUS.items():
                if key.upper() in fournisseur.upper():
                    return value
            return fournisseur[:35]
        # Fallback
        parts = intitule.replace('VIR SEPA', '').strip().split()
        if parts:
            fournisseur = ' '.join(parts[:2])
            for key, value in FOURNISSEURS_CONNUS.items():
                if key.upper() in fournisseur.upper():
                    return value
            return fournisseur[:35]
    
    # ========== VIR INST ==========
    if 'VIR INST' in intitule:
        match = re.search(r'VIR INST\s+([A-Za-z][A-Za-z0-9\s\.\'\-&]+?)(?:\s+(?:Pour|pour|P\d|\d{4,})|\s+-\s+|\s*$)', intitule, re.IGNORECASE)
        if match:
            fournisseur = match.group(1).strip()
            for key, value in FOURNISSEURS_CONNUS.items():
                if key.upper() in fournisseur.upper():
                    return value
            return fournisseur[:35]
        parts = intitule.replace('VIR INST', '').strip().split()
        if parts:
            fournisseur = ' '.join(parts[:2])
            for key, value in FOURNISSEURS_CONNUS.items():
                if key.upper() in fournisseur.upper():
                    return value
            return fournisseur[:35]
    
    # ========== VIR PERM ==========
    if 'VIR PERM' in intitule:
        match = re.search(r'VIR PERM\s+([A-Za-z][A-Za-z0-9\s\.\'\-&]+?)(?:\s+(?:LOCATION|Pour|pour|P\d|\d{4,})|\s+-\s+|\s*$)', intitule, re.IGNORECASE)
        if match:
            fournisseur = match.group(1).strip()
            return fournisseur[:35]
    
    # ========== PRLV SEPA ==========
    if 'PRLV SEPA' in intitule:
        match = re.search(r'PRLV SEPA\s+([A-Za-z][A-Za-z0-9\s\.\'\-&]+?)(?:\s+(?:Votre|Maintenance|COTISATION|THERMALE|RELAIS|I\d|FA\d|F\d|\d{6,}|PRELEVEMENT|FACTURE|UR\s|XX|FR\d|Prelevement|Ref|\+\+)|\s+-\s+|\s*$)', intitule, re.IGNORECASE)
        if match:
            fournisseur = match.group(1).strip()
            # Vérifier dans le dictionnaire
            for key, value in FOURNISSEURS_CONNUS.items():
                if key.upper() in fournisseur.upper():
                    return value
            return fournisseur[:35]
        # Fallback
        parts = intitule.replace('PRLV SEPA', '').strip().split()
        if parts:
            fournisseur = ' '.join(parts[:2])
            for key, value in FOURNISSEURS_CONNUS.items():
                if key.upper() in fournisseur.upper():
                    return value
            return fournisseur[:35]
    
    # ========== REMBOURSEMENT PRET ==========
    if 'ECH PRET' in intitule or 'REMBOURSEMENT PRET' in intitule_upper:
        return "Banque (Remboursement prêt)"
    
    # ========== FRAIS BANCAIRES ==========
    if 'ABON' in intitule_upper and ('CYBERPLUS' in intitule_upper or 'BANQUE' in intitule_upper):
        return "Frais bancaires"
    
    if 'COM CB' in intitule_upper or 'COMM.' in intitule_upper:
        return "Frais bancaires"
    
    if 'ARRETE COMPTE' in intitule_upper:
        return "Frais bancaires"
    
    if 'FRAIS PAI' in intitule_upper or 'FRAIS CB' in intitule_upper:
        return "Frais bancaires"
    
    # ========== PAIEMENTS CB ==========
    if 'CB' in intitule_upper and '****' in intitule:
        return "Paiement CB (divers)"
    
    if 'CARTE PAIEMENTS' in intitule_upper or 'CYBER+' in intitule_upper:
        return "Paiement CB (divers)"
    
    # ========== DEFAULT ==========
    # Par défaut, prendre les premiers mots significatifs
    mots = [m for m in intitule.split()[:4] if len(m) > 2 and not m.isdigit()]
    if mots:
        result = ' '.join(mots[:2])
        # Vérifier une dernière fois dans le dictionnaire
        for key, value in FOURNISSEURS_CONNUS.items():
            if key.upper() in result.upper():
                return value
        return result[:35]
    
    return "Non spécifié"

# ============================================================================
# FONCTION DE CATÉGORISATION (utilise la catégorie existante)
# ============================================================================
def categoriser_depense(row):
    """Catégorise une dépense en utilisant la catégorie existante"""
    categorie = str(row['Categorie']) if pd.notna(row['Categorie']) else ""
    intitule = str(row['Intitule']) if pd.notna(row['Intitule']) else ""
    
    # Utiliser la catégorie existante avec icônes
    mapping_categories = {
        'BOUCHERIE': '🥩 Boucherie',
        'EPICERIE': '🛒 Épicerie',
        'BOF': '🧀 Produits laitiers (BOF)',
        'LEGUMES/FRUITS': '🥕 Fruits & Légumes',
        'CAVE': '🍷 Cave & Boissons',
        'BOULANGERIE DESSERTS': '🥐 Boulangerie & Desserts',
        'GLACES': '🍨 Glaces',
        'LAVAGE LINGE': '🧺 Blanchisserie',
        'PRODUITS ENTRETIEN': '🧹 Produits d\'entretien',
        'PETIT MATERIEL': '🔧 Petit matériel',
        'SOUS-TRAITANT': '👷 Sous-traitance',
        'MAINTENANCE': '🔧 Maintenance',
        'REPARATION': '🛠️ Réparations',
        'IVESTISSEMENT BTP': '🏗️ Investissements BTP',
        'LOCATION': '📋 Locations & Abonnements',
        'TELEPHONE INTERNET': '📞 Téléphone & Internet',
        'BURAUTIQUE': '🖨️ Bureautique',
        'PUB/COM': '📢 Publicité & Communication',
        'COMMISSION': '💳 Commissions',
        'HONORAIRE': '💼 Honoraires',
        'ASSURANCE': '🛡️ Assurances',
        'EAU': '💧 Eau',
        'ELECTRICITÉ & GAZ': '⚡ Électricité & Gaz',
        'Electricité & Gaz': '⚡ Électricité & Gaz',
        'SALAIRE': '👥 Salaires',
        'Salaire': '👥 Salaires',
        'URSSAF': '🏛️ URSSAF',
        'Urssaf': '🏛️ URSSAF',
        'RETRAITES': '🏛️ Retraites',
        'Retraites': '🏛️ Retraites',
        'RECRUTEMENT': '👥 Recrutement',
        'Recrutement': '👥 Recrutement',
        'FRAIS BANCAIRE': '🏦 Frais bancaires',
        'frais bancaire': '🏦 Frais bancaires',
        'FRAIS BANQUE': '🏦 Frais bancaires',
        'PRET MARZE 2030 150 KE': '🏦 Prêt MARZE',
        'PRET CA 2030 518 KE': '🏦 Prêt CA',
        'BPI 2031 350 KE': '🏦 Prêt BPI',
        'TVA': '📊 TVA',
        'TF': '📋 Taxe Foncière',
        'TAXES SEJOUR': '📋 Taxes de séjour',
        'SACEM': '🎵 SACEM',
        'AFFRANCHISSEMENT': '📮 Affranchissement',
        'FRAIS REPRESENTATION': '🎁 Frais de représentation',
        'REBOURSEMENT CLIENT TROP VERSE': '💸 Remboursement client',
        'A CATÉGORISER': '❓ À catégoriser',
        'A Catégoriser': '❓ À catégoriser',
        'AUTRE': '📦 Autres',
        'Autre': '📦 Autres',
    }
    
    # Chercher dans le mapping
    for key, value in mapping_categories.items():
        if key.upper() == categorie.upper():
            return value
    
    # Cas spéciaux basés sur l'intitulé
    if 'AMAZON' in intitule.upper():
        return '🛒 Achats Amazon'
    if 'GOOGLE' in intitule.upper() and 'ADS' in intitule.upper():
        return '📢 Publicité & Communication'
    if 'LA POSTE' in intitule.upper():
        return '📮 Affranchissement'
    
    # Si catégorie existante mais pas dans le mapping
    if categorie and categorie != 'nan':
        return f"📦 {categorie}"
    
    return "❓ À catégoriser"

# ============================================================================
# APPLIQUER LES FONCTIONS
# ============================================================================
print("🏷️  Extraction des fournisseurs et catégorisation...")
df['Fournisseur'] = df.apply(lambda row: extraire_fournisseur(row['Intitule']), axis=1)
df['Categorie_finale'] = df.apply(categoriser_depense, axis=1)

# ============================================================================
# PRÉPARATION DES DONNÉES POUR LE TABLEAU DE BORD
# ============================================================================
print("📊 Préparation des analyses...")

# Analyse par catégorie
par_categorie = df.groupby('Categorie_finale').agg({
    'Montant_TTC': ['sum', 'count', 'mean', 'min', 'max']
}).round(2)
par_categorie.columns = ['Total TTC (€)', 'Nb commandes', 'Moyenne (€)', 'Min (€)', 'Max (€)']
par_categorie = par_categorie.sort_values('Total TTC (€)', ascending=False)
par_categorie['% du total'] = (par_categorie['Total TTC (€)'] / par_categorie['Total TTC (€)'].sum() * 100).round(1)
par_categorie = par_categorie.reset_index()
par_categorie.columns = ['Catégorie', 'Total TTC (€)', 'Nb commandes', 'Moyenne (€)', 'Min (€)', 'Max (€)', '% du total']

# Analyse par fournisseur (Top 40)
par_fournisseur = df.groupby('Fournisseur').agg({
    'Montant_TTC': ['sum', 'count', 'mean']
}).round(2)
par_fournisseur.columns = ['Total TTC (€)', 'Nb commandes', 'Moyenne (€)']
par_fournisseur = par_fournisseur.sort_values('Total TTC (€)', ascending=False).head(40)
par_fournisseur['% du total'] = (par_fournisseur['Total TTC (€)'] / df['Montant_TTC'].sum() * 100).round(1)
par_fournisseur = par_fournisseur.reset_index()

# Données détaillées
detail = df[['Categorie_finale', 'Fournisseur', 'Montant_TTC', 'Intitule']].copy()
detail.columns = ['Catégorie', 'Fournisseur', 'Montant TTC (€)', 'Intitulé']
detail = detail.sort_values('Montant TTC (€)', ascending=False)

# KPIs
total_depenses = df['Montant_TTC'].sum()
nb_operations = len(df)
moyenne_par_operation = df['Montant_TTC'].mean()
nb_categories = df['Categorie_finale'].nunique()
nb_fournisseurs = df['Fournisseur'].nunique()
top_categorie = par_categorie.iloc[0]['Catégorie']
top_categorie_montant = par_categorie.iloc[0]['Total TTC (€)']
top_fournisseur = par_fournisseur.iloc[0]['Fournisseur']
top_fournisseur_montant = par_fournisseur.iloc[0]['Total TTC (€)']

# ============================================================================
# CRÉATION DU FICHIER EXCEL
# ============================================================================
print("📝 Création du tableau de bord Excel...")

wb = Workbook()

# Styles
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
title_font = Font(bold=True, size=24, color="1F4E79")
subtitle_font = Font(bold=True, size=14, color="1F4E79")
kpi_font = Font(bold=True, size=18, color="2E7D32")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center')
money_format = '#,##0.00 €'
percent_format = '0.0%'

# ============================================================================
# FEUILLE 1: TABLEAU DE BORD PRINCIPAL
# ============================================================================
ws_dashboard = wb.active
ws_dashboard.title = "📊 Tableau de bord"

# Titre
ws_dashboard.merge_cells('B2:I2')
ws_dashboard['B2'] = "🏨 TABLEAU DE BORD DES DÉPENSES 2025"
ws_dashboard['B2'].font = title_font
ws_dashboard['B2'].alignment = center_align

ws_dashboard.merge_cells('B3:I3')
ws_dashboard['B3'] = "Grand Hôtel des Bains"
ws_dashboard['B3'].font = subtitle_font
ws_dashboard['B3'].alignment = center_align

# KPIs principaux
kpi_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
kpi_fill2 = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
kpi_fill3 = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

# Ligne KPI 1
ws_dashboard.merge_cells('B5:C6')
ws_dashboard['B5'] = "💰 TOTAL DÉPENSES"
ws_dashboard['B5'].font = Font(bold=True, size=11)
ws_dashboard['B5'].alignment = center_align
ws_dashboard['B5'].fill = kpi_fill

ws_dashboard.merge_cells('D5:E6')
ws_dashboard['D5'] = f"{total_depenses:,.2f} €"
ws_dashboard['D5'].font = kpi_font
ws_dashboard['D5'].alignment = center_align
ws_dashboard['D5'].fill = kpi_fill

ws_dashboard.merge_cells('F5:G6')
ws_dashboard['F5'] = "📋 NB OPÉRATIONS"
ws_dashboard['F5'].font = Font(bold=True, size=11)
ws_dashboard['F5'].alignment = center_align
ws_dashboard['F5'].fill = kpi_fill2

ws_dashboard.merge_cells('H5:I6')
ws_dashboard['H5'] = nb_operations
ws_dashboard['H5'].font = kpi_font
ws_dashboard['H5'].alignment = center_align
ws_dashboard['H5'].fill = kpi_fill2

# Ligne KPI 2
ws_dashboard.merge_cells('B8:C9')
ws_dashboard['B8'] = "📊 MOYENNE/OPÉRATION"
ws_dashboard['B8'].font = Font(bold=True, size=11)
ws_dashboard['B8'].alignment = center_align
ws_dashboard['B8'].fill = kpi_fill3

ws_dashboard.merge_cells('D8:E9')
ws_dashboard['D8'] = f"{moyenne_par_operation:,.2f} €"
ws_dashboard['D8'].font = kpi_font
ws_dashboard['D8'].alignment = center_align
ws_dashboard['D8'].fill = kpi_fill3

ws_dashboard.merge_cells('F8:G9')
ws_dashboard['F8'] = "🏢 NB FOURNISSEURS"
ws_dashboard['F8'].font = Font(bold=True, size=11)
ws_dashboard['F8'].alignment = center_align
ws_dashboard['F8'].fill = kpi_fill

ws_dashboard.merge_cells('H8:I9')
ws_dashboard['H8'] = nb_fournisseurs
ws_dashboard['H8'].font = kpi_font
ws_dashboard['H8'].alignment = center_align
ws_dashboard['H8'].fill = kpi_fill

# Top Catégorie et Fournisseur
ws_dashboard.merge_cells('B11:E12')
ws_dashboard['B11'] = f"🏆 Top Catégorie: {top_categorie}"
ws_dashboard['B11'].font = Font(bold=True, size=12, color="D84315")
ws_dashboard['B11'].alignment = left_align

ws_dashboard.merge_cells('B13:E13')
ws_dashboard['B13'] = f"    → {top_categorie_montant:,.2f} €"
ws_dashboard['B13'].font = Font(size=11)

ws_dashboard.merge_cells('F11:I12')
ws_dashboard['F11'] = f"🏆 Top Fournisseur: {top_fournisseur}"
ws_dashboard['F11'].font = Font(bold=True, size=12, color="D84315")
ws_dashboard['F11'].alignment = left_align

ws_dashboard.merge_cells('F13:I13')
ws_dashboard['F13'] = f"    → {top_fournisseur_montant:,.2f} €"
ws_dashboard['F13'].font = Font(size=11)

# Mini tableau des top 10 catégories
ws_dashboard['B16'] = "📊 TOP 10 CATÉGORIES"
ws_dashboard['B16'].font = subtitle_font

headers_cat = ['Catégorie', 'Total TTC (€)', 'Nb cmd', '% total']
for col, header in enumerate(headers_cat, start=2):
    cell = ws_dashboard.cell(row=17, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border

for i, row in par_categorie.head(10).iterrows():
    row_num = 18 + i
    ws_dashboard.cell(row=row_num, column=2, value=row['Catégorie']).border = border
    ws_dashboard.cell(row=row_num, column=3, value=row['Total TTC (€)']).border = border
    ws_dashboard.cell(row=row_num, column=3).number_format = money_format
    ws_dashboard.cell(row=row_num, column=4, value=row['Nb commandes']).border = border
    ws_dashboard.cell(row=row_num, column=5, value=row['% du total']/100).border = border
    ws_dashboard.cell(row=row_num, column=5).number_format = percent_format

# Mini tableau des top 10 fournisseurs
ws_dashboard['G16'] = "🏢 TOP 10 FOURNISSEURS"
ws_dashboard['G16'].font = subtitle_font

headers_four = ['Fournisseur', 'Total TTC (€)', '% total']
for col, header in enumerate(headers_four, start=7):
    cell = ws_dashboard.cell(row=17, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border

for i, row in par_fournisseur.head(10).iterrows():
    row_num = 18 + i
    ws_dashboard.cell(row=row_num, column=7, value=row['Fournisseur']).border = border
    ws_dashboard.cell(row=row_num, column=8, value=row['Total TTC (€)']).border = border
    ws_dashboard.cell(row=row_num, column=8).number_format = money_format
    ws_dashboard.cell(row=row_num, column=9, value=row['% du total']/100).border = border
    ws_dashboard.cell(row=row_num, column=9).number_format = percent_format

# Ajuster les largeurs de colonnes
ws_dashboard.column_dimensions['A'].width = 3
ws_dashboard.column_dimensions['B'].width = 30
ws_dashboard.column_dimensions['C'].width = 15
ws_dashboard.column_dimensions['D'].width = 15
ws_dashboard.column_dimensions['E'].width = 12
ws_dashboard.column_dimensions['F'].width = 5
ws_dashboard.column_dimensions['G'].width = 30
ws_dashboard.column_dimensions['H'].width = 15
ws_dashboard.column_dimensions['I'].width = 12

# ============================================================================
# FEUILLE 2: ANALYSE PAR CATÉGORIE
# ============================================================================
ws_cat = wb.create_sheet("📂 Par Catégorie")

ws_cat['A1'] = "📂 ANALYSE PAR CATÉGORIE"
ws_cat['A1'].font = title_font

# En-têtes
headers = ['Catégorie', 'Total TTC (€)', 'Nb commandes', 'Moyenne (€)', 'Min (€)', 'Max (€)', '% du total']
for col, header in enumerate(headers, start=1):
    cell = ws_cat.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border

# Données
for i, row in par_categorie.iterrows():
    row_num = 4 + i
    ws_cat.cell(row=row_num, column=1, value=row['Catégorie']).border = border
    ws_cat.cell(row=row_num, column=2, value=row['Total TTC (€)']).border = border
    ws_cat.cell(row=row_num, column=2).number_format = money_format
    ws_cat.cell(row=row_num, column=3, value=row['Nb commandes']).border = border
    ws_cat.cell(row=row_num, column=4, value=row['Moyenne (€)']).border = border
    ws_cat.cell(row=row_num, column=4).number_format = money_format
    ws_cat.cell(row=row_num, column=5, value=row['Min (€)']).border = border
    ws_cat.cell(row=row_num, column=5).number_format = money_format
    ws_cat.cell(row=row_num, column=6, value=row['Max (€)']).border = border
    ws_cat.cell(row=row_num, column=6).number_format = money_format
    ws_cat.cell(row=row_num, column=7, value=row['% du total']/100).border = border
    ws_cat.cell(row=row_num, column=7).number_format = percent_format

# Ligne de total
total_row = 4 + len(par_categorie)
ws_cat.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
ws_cat.cell(row=total_row, column=1).fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
ws_cat.cell(row=total_row, column=2, value=par_categorie['Total TTC (€)'].sum())
ws_cat.cell(row=total_row, column=2).font = Font(bold=True)
ws_cat.cell(row=total_row, column=2).number_format = money_format
ws_cat.cell(row=total_row, column=2).fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
ws_cat.cell(row=total_row, column=3, value=par_categorie['Nb commandes'].sum())
ws_cat.cell(row=total_row, column=3).font = Font(bold=True)
ws_cat.cell(row=total_row, column=3).fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")

# Graphique en barres pour les catégories
chart_bar = BarChart()
chart_bar.type = "bar"
chart_bar.style = 10
chart_bar.title = "Dépenses par Catégorie"
chart_bar.y_axis.title = "Catégorie"
chart_bar.x_axis.title = "Montant (€)"

data = Reference(ws_cat, min_col=2, min_row=3, max_row=3+len(par_categorie))
cats = Reference(ws_cat, min_col=1, min_row=4, max_row=3+len(par_categorie))
chart_bar.add_data(data, titles_from_data=True)
chart_bar.set_categories(cats)
chart_bar.shape = 4
chart_bar.width = 20
chart_bar.height = 15

ws_cat.add_chart(chart_bar, "I3")

# Ajuster les largeurs
ws_cat.column_dimensions['A'].width = 35
ws_cat.column_dimensions['B'].width = 15
ws_cat.column_dimensions['C'].width = 14
ws_cat.column_dimensions['D'].width = 14
ws_cat.column_dimensions['E'].width = 12
ws_cat.column_dimensions['F'].width = 12
ws_cat.column_dimensions['G'].width = 12

# ============================================================================
# FEUILLE 3: ANALYSE PAR FOURNISSEUR
# ============================================================================
ws_four = wb.create_sheet("🏢 Par Fournisseur")

ws_four['A1'] = "🏢 TOP 40 FOURNISSEURS"
ws_four['A1'].font = title_font

# En-têtes
headers = ['Fournisseur', 'Total TTC (€)', 'Nb commandes', 'Moyenne (€)', '% du total']
for col, header in enumerate(headers, start=1):
    cell = ws_four.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border

# Données
for i, row in par_fournisseur.iterrows():
    row_num = 4 + i
    ws_four.cell(row=row_num, column=1, value=row['Fournisseur']).border = border
    ws_four.cell(row=row_num, column=2, value=row['Total TTC (€)']).border = border
    ws_four.cell(row=row_num, column=2).number_format = money_format
    ws_four.cell(row=row_num, column=3, value=row['Nb commandes']).border = border
    ws_four.cell(row=row_num, column=4, value=row['Moyenne (€)']).border = border
    ws_four.cell(row=row_num, column=4).number_format = money_format
    ws_four.cell(row=row_num, column=5, value=row['% du total']/100).border = border
    ws_four.cell(row=row_num, column=5).number_format = percent_format

# Graphique en barres pour les fournisseurs (top 15)
chart_four = BarChart()
chart_four.type = "bar"
chart_four.style = 11
chart_four.title = "Top 15 Fournisseurs"
chart_four.y_axis.title = "Fournisseur"
chart_four.x_axis.title = "Montant (€)"

data = Reference(ws_four, min_col=2, min_row=3, max_row=18)
cats = Reference(ws_four, min_col=1, min_row=4, max_row=18)
chart_four.add_data(data, titles_from_data=True)
chart_four.set_categories(cats)
chart_four.width = 18
chart_four.height = 12

ws_four.add_chart(chart_four, "G3")

# Ajuster les largeurs
ws_four.column_dimensions['A'].width = 35
ws_four.column_dimensions['B'].width = 15
ws_four.column_dimensions['C'].width = 14
ws_four.column_dimensions['D'].width = 14
ws_four.column_dimensions['E'].width = 12

# ============================================================================
# FEUILLE 4: DONNÉES DÉTAILLÉES
# ============================================================================
ws_detail = wb.create_sheet("📋 Détail")

ws_detail['A1'] = "📋 DÉTAIL DE TOUTES LES DÉPENSES"
ws_detail['A1'].font = title_font

# En-têtes
headers = ['Catégorie', 'Fournisseur', 'Montant TTC (€)', 'Intitulé']
for col, header in enumerate(headers, start=1):
    cell = ws_detail.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border

# Données
for i, (idx, row) in enumerate(detail.iterrows()):
    row_num = 4 + i
    ws_detail.cell(row=row_num, column=1, value=row['Catégorie']).border = border
    ws_detail.cell(row=row_num, column=2, value=row['Fournisseur']).border = border
    ws_detail.cell(row=row_num, column=3, value=row['Montant TTC (€)']).border = border
    ws_detail.cell(row=row_num, column=3).number_format = money_format
    intitule = str(row['Intitulé'])[:80] if pd.notna(row['Intitulé']) else ""
    ws_detail.cell(row=row_num, column=4, value=intitule).border = border

# Ajuster les largeurs
ws_detail.column_dimensions['A'].width = 32
ws_detail.column_dimensions['B'].width = 35
ws_detail.column_dimensions['C'].width = 15
ws_detail.column_dimensions['D'].width = 80

# ============================================================================
# SAUVEGARDE
# ============================================================================
fichier_sortie = "TABLEAU_DE_BORD_DEPENSES_2025.xlsx"
wb.save(fichier_sortie)

print(f"\n{'='*80}")
print(f"✅ TABLEAU DE BORD CRÉÉ: {fichier_sortie}")
print(f"{'='*80}")
print(f"\n📊 Contenu du fichier:")
print(f"   1️⃣  📊 Tableau de bord - Vue d'ensemble avec KPIs")
print(f"   2️⃣  📂 Par Catégorie - {len(par_categorie)} catégories + graphique")
print(f"   3️⃣  🏢 Par Fournisseur - Top 40 fournisseurs + graphique")
print(f"   4️⃣  📋 Détail - Toutes les {len(detail)} opérations")

print(f"\n📈 RÉSUMÉ:")
print(f"   💰 Total dépenses: {total_depenses:,.2f} €")
print(f"   📋 Nb opérations: {nb_operations}")
print(f"   🏢 Nb fournisseurs: {nb_fournisseurs}")
print(f"   🏷️  Nb catégories: {nb_categories}")

print(f"\n🔍 APERÇU DES FOURNISSEURS EXTRAITS:")
for i, row in par_fournisseur.head(15).iterrows():
    print(f"   • {row['Fournisseur']}: {row['Total TTC (€)']:,.2f} € ({row['Nb commandes']} cmd)")

print(f"\n💡 Ouvrez le fichier Excel pour voir les graphiques!")
